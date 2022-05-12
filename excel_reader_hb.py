#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Dec 13 16:21:39 2020

@author: mirac
"""


from openpyxl import Workbook
import openpyxl
import os



def addsheet(cat_list,kat):
    sheetname="hb_kat_list/"+str(cat_list[0])+".xlsx"
    
    if os.path.exists(sheetname) == False:
        #dosya yoksa oluşturulur
        book = Workbook()
        sheet = book.active
        sheet["A1"]='Sektör'
        sheet["B1"]='Üst Kategori'
        sheet["C1"]='Kategori'
        sheet["D1"]='Alt Kategori'
        sheet["E1"]='Alt Kategori'
        sheet["F1"]='Alt Kategori'
        sheet["G1"]='Alt Kategori'
        sheet["H1"]='Alt Kategori'
        sheet["I1"]='Alt Kategori'
        sheet["J1"]='Alt Kategori'
        
        #dosya oluşturulurken sayfa başlığını şekildeki gibi ekleyip bir alttaki kod satırıyla kaydediyoruz.
        book.save(sheetname)
        print("Dosya oluşturuldu..\nFile created..")
    
    sheet1 = openpyxl.load_workbook(filename = sheetname)
    sheet=sheet1.active
    #yukarıdaki satırlarda kayıt edildilten sonra kapanan excel sayfamızı yeniden açıp yazmak için aktif hale getirdik.
    
    
    
    #yukarıdaki kod satırlarında entity içerisinden alınan bilgileri değişkenler içinde tutuyoruz.
    
    #max_rowi=sheet.max_row
    #datasheet içindeki satır sayıları kaç adet kayıt var onun sayısını döndürür
    #+1 demek ise bizim kaydımızı yazmak istediğimiz satır
    #print(cat_list)
    #print(type(cat_list))

    sheet["A"+str(sheet.max_row+1)]=cat_list[0] if str(cat_list[0]) not in kat else ' '
    sheet["B"+str(sheet.max_row)]=cat_list[1] if len(cat_list)>1 and type(cat_list[1]) != list and cat_list[1] not in kat else ""
    sheet["C"+str(sheet.max_row)]=cat_list[2] if len(cat_list)>2 and type(cat_list[2]) != list and cat_list[2] not in kat else ""
    sheet["D"+str(sheet.max_row)]=cat_list[3] if len(cat_list)>3 and type(cat_list[3]) != list and cat_list[3] not in kat else ""
    sheet["E"+str(sheet.max_row)]=cat_list[4] if len(cat_list)>4 and type(cat_list[4]) != list and cat_list[4] not in kat else ""
    sheet["F"+str(sheet.max_row)]=cat_list[5] if len(cat_list)>5 and type(cat_list[5]) != list and cat_list[5] not in kat else ""
    sheet["G"+str(sheet.max_row)]=cat_list[6] if len(cat_list)>6 and type(cat_list[6]) != list and cat_list[6] not in kat else ""
    sheet["H"+str(sheet.max_row)]=cat_list[7] if len(cat_list)>7 and type(cat_list[7]) != list and cat_list[7] not in kat else ""
    sheet["I"+str(sheet.max_row)]=cat_list[8] if len(cat_list)>8 and type(cat_list[8]) != list and cat_list[8] not in kat else ""
    count=sheet.max_row
    for i in cat_list[-1]:
        sheet["J"+str(count)]=i
        count=count+1
        
    
    #ilgili hücrenin herbirine gelecek olan değişkenler içinde tuttuğumuz kullanıcıdan alınan bilgileri yazdık
    
    
    sheet1.save(sheetname)
    #save diyerek kaydedip kapattık
    print("Form gönderildi..")
