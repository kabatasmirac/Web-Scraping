#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Dec 13 16:19:54 2020

@author: mirac
"""



from openpyxl import Workbook
import openpyxl
import os



def addsheet(cat_list):
    sheetname="opt_attr.xlsx"
    
    if os.path.exists(sheetname) == False:
        #dosya yoksa oluşturulur
        book = Workbook()
        sheet = book.active
        sheet["A1"]='Kategori'
        sheet["B1"]='Özellik Adı'
        sheet["C1"]='Özellik Seçenekleri'        
        #dosya oluşturulurken sayfa başlığını şekildeki gibi ekleyip bir alttaki kod satırıyla kaydediyoruz.
        book.save(sheetname)
        print("Dosya oluşturuldu..\nFile created..")
    
    sheet1 = openpyxl.load_workbook(filename = sheetname)
    sheet=sheet1.active
    #yukarıdaki satırlarda kayıt edildilten sonra kapanan excel sayfamızı yeniden açıp yazmak için aktif hale getirdik.
    
    
    
    #yukarıdaki kod satırlarında entity içerisinden alınan bilgileri değişkenler içinde tutuyoruz.
    
    #max_rowi=sheet.max_row
    #datasheet içindeki satır sayıları kaç adet kayıt var onun sayısını döndürür
    #+1 demek ise bizim kaydımızı yazmak istediğimiz satır
    #print(cat_list)
    #print(type(cat_list))
    sheet["A"+str(sheet.max_row+1)]=cat_list[0]
    sheet["B"+str(sheet.max_row)]=cat_list[1] if len(cat_list)>1 and type(cat_list[1]) != list else "NULL"
    sheet["C"+str(sheet.max_row)]=cat_list[2] if len(cat_list)>2 and type(cat_list[2]) != list else "NULL"
        
    
    #ilgili hücrenin herbirine gelecek olan değişkenler içinde tuttuğumuz kullanıcıdan alınan bilgileri yazdık
    
    
    sheet1.save(sheetname)
    #save diyerek kaydedip kapattık
    print("Form gönderildi..")
