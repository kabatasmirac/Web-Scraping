#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Dec 12 16:48:59 2020

@author: mirac
"""

from openpyxl import Workbook
import openpyxl
import os



def addsheet(cat_list,kat):
    
    sheetname="trendyol_kategori_subat.xlsx"
    if os.path.exists(sheetname) == False:
        #dosya yoksa oluşturulur
        book = Workbook()
        sheet = book.active
        sheet["A1"]='Sektör'
        sheet["B1"]='Üst Kategori'
        sheet["C1"]='Kategori'
        sheet["D1"]='Alt Kategori'
        sheet["E1"]='Özellikler'
        
        #dosya oluşturulurken sayfa başlığını şekildeki gibi ekleyip bir alttaki kod satırıyla kaydediyoruz.
        book.save(sheetname)
        print("Dosya oluşturuldu..\nFile created..")
    
    sheet1 = openpyxl.load_workbook(filename = sheetname)
    sheet=sheet1.active
        
    
    #yukarıdaki satırlarda kayıt edildilten sonra kapanan excel sayfamızı yeniden açıp yazmak için aktif hale getirdik.
    
    
    
    #yukarıdaki kod satırlarında entity içerisinden alınan bilgileri değişkenler içinde tutuyoruz.
    
    #max_rowi=sheet.max_row
    #datasheet içindeki satır sayıları kaç adet kayıt var onun sayısını döndürür
    #+1 demek ise bizim kaydımızı yazmak istediğimiz satır
    #print(cat_list)
    #print(type(cat_list))
    if str(cat_list[0]).strip() not in kat:
        sheet["A"+str(sheet.max_row+1)]=cat_list[0]
        sheet["B"+str(sheet.max_row)]=cat_list[1] if len(cat_list)>1 and type(cat_list[1]) != list else "NULL"
        sheet["C"+str(sheet.max_row)]=cat_list[2] if len(cat_list)>2 and type(cat_list[2]) != list else "NULL"
        sheet["D"+str(sheet.max_row)]=cat_list[3] if len(cat_list)>3 and type(cat_list[3]) != list else "NULL"
        count=sheet.max_row
    else:
        sheet["B"+str(sheet.max_row)]=cat_list[1] if len(cat_list)>1 and type(cat_list[1]) != list else "NULL"
        sheet["C"+str(sheet.max_row)]=cat_list[2] if len(cat_list)>2 and type(cat_list[2]) != list else "NULL"
        sheet["D"+str(sheet.max_row)]=cat_list[3] if len(cat_list)>3 and type(cat_list[3]) != list else "NULL"
        count=sheet.max_row
    
    for i in cat_list[-1]:
        sheet["E"+str(count)]=i
        count=count+1
        
    
    #ilgili hücrenin herbirine gelecek olan değişkenler içinde tuttuğumuz kullanıcıdan alınan bilgileri yazdık
    
    
    sheet1.save(sheetname)
    #save diyerek kaydedip kapattık
    print("Form gönderildi..")
