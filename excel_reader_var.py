#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Dec 21 20:57:25 2020

@author: mirac
"""

from openpyxl import Workbook
import openpyxl
import os



def addsheet(cat_list):
    
    sheetname="varyant_son/trendyol_varient_tumu123.xlsx"
    if os.path.exists(sheetname) == False:
        #dosya yoksa oluşturulur
        book = Workbook()
        sheet = book.active
        sheet["A1"]='Sektör'
        sheet["B1"]='Özellik'
        
        sheet["C1"]='Varyant mı?'
        sheet["D1"]='Varyant Özellikleri'
        
        #dosya oluşturulurken sayfa başlığını şekildeki gibi ekleyip bir alttaki kod satırıyla kaydediyoruz.
        book.save(sheetname)
        print("Dosya oluşturuldu..\nFile created..")
    
    sheet1 = openpyxl.load_workbook(filename = sheetname)
    sheet=sheet1.active
        
    
    #yukarıdaki satırlarda kayıt edildilten sonra kapanan excel sayfamızı yeniden açıp yazmak için aktif hale getirdik.
    
    
    
    #yukarıdaki kod satırlarında entity içerisinden alınan bilgileri değişkenler içinde tutuyoruz.
    
    #max_rowi=sheet.max_row
    #datasheet içindeki satır sayıları kaç adet kayıt var onun sayısını döndürür
    #+1 demek ise bizim kaydımızı yazmak istediğimiz satır
    #print(cat_list)
    #print(type(cat_list))
    

    sheet["A"+str(sheet.max_row+1)]=cat_list[0]
    sheet["B"+str(sheet.max_row)]=cat_list[1] 
    
    sheet["C"+str(sheet.max_row)]= 'Evet' if cat_list[2] is True else 'Hayır'
    
    sheet["D"+str(sheet.max_row)]=str(cat_list[3]) 
        
        
        
    
    #ilgili hücrenin herbirine gelecek olan değişkenler içinde tuttuğumuz kullanıcıdan alınan bilgileri yazdık
    
    
    sheet1.save(sheetname)
    #save diyerek kaydedip kapattık
    print("Form gönderildi..")
